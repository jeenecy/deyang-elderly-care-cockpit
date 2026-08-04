import openpyxl, json
path = "C:/Users/宋志超/Downloads/卫生专网签约情况统计表（在线）.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
print("SHEETS:", wb.sheetnames)
out = {}
for ws in wb.worksheets:
    rows = list(ws.iter_rows(values_only=True))
    # strip trailing fully-empty rows
    nonempty = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
    print(f"\n=== SHEET: {ws.title} | dims={ws.dimensions} | rows={len(rows)} nonempty={len(nonempty)} ===")
    for i, r in enumerate(nonempty[:60]):
        print(i, [("" if c is None else str(c)) for c in r])
    out[ws.title] = [[("" if c is None else str(c)) for c in r] for r in nonempty[:200]]
with open("C:/Users/宋志超/WorkBuddy/data-elderly care/scripts/xlsx_dump.json", "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print("\nDUMPED -> scripts/xlsx_dump.json")
